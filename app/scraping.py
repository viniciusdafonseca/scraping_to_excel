import re

from bs4 import BeautifulSoup
from loguru import logger
from openpyxl import load_workbook

from app.utils.async_patchright import Patchright
from app.utils.captcha import AntiCaptcha


class Scraping:
    def __init__(self, cpf_cnpj: str) -> None:
        self._cpf_cnpj = re.sub(r"\D", "", cpf_cnpj)
        self.playwright = Patchright(self._cpf_cnpj, headless=True)
        self._url = (
            "https://www.gp.srv.br/tributario/querencia/portal_serv_servico?7,61"
        )
        self._captcha = AntiCaptcha()

    async def start(self) -> None:
        self.context = await self.playwright.start()
        self.page = await self.context.new_page()

    async def finish(self) -> None:
        await self.playwright.finish()

    async def _solve_captcha(self) -> None:
        """
        Resolve um CAPTCHA capturando sua imagem, processa via serviço externo
        e preenche o resultado no campo adequado da página web.

        :return: None
        :rtype: NoneType
        """
        img = await self.page.locator(
            'xpath=//*[@id="W0005W0006TXTKAPTCHA"]/img'
        ).screenshot()
        resolved_captcha = await self._captcha.image_captcha(img)
        await self.page.fill("#W0005W0006vVALOR_IMAGEM", resolved_captcha)

    async def _login_sistema(self) -> BeautifulSoup:
        """
        Faz login no sistema e retorna o conteúdo da página no formato do BeautifulSoup.
        :return: Conteúdo HTML parseado.
        :rtype: BeautifulSoup
        """
        await self.page.goto(self._url)

        await self.page.fill("#W0005W0006vCAMPO_00010001", self._cpf_cnpj)
        await self._solve_captcha()
        logger.info("Formulário preenchido com sucesso!")

        await self.page.click("#W0005W0006BTN_CONSULTAR2")
        logger.info("Botão de consulta pressionado!")
        await self.page.wait_for_selector("text=IPTU - 922704/2025", timeout=120000)
        logger.info("Página carregada!")
        return BeautifulSoup(await self.page.content(), "html.parser")

    async def run(self) -> None:
        """
        Realiza a extração de dados de imóveis do sistema web e salva em planilha.

        Este método faz login no sistema, extrai informações de imóveis (inscrição,
        bairro, quadra e lote) e salva os dados em uma planilha Excel.

        :return: None
        """
        infos_lista = []

        soup = await self._login_sistema()
        div_inscricoes = soup.find_all("div", {"class": "card-deb"})
        for div in div_inscricoes:
            inscricao = re.search(r"Imóvel (\d*)", div.text, re.M)[1]
            bairro = re.search(r"Bairro (.*?) Quadra", div.text, re.M)[1]
            quadra = re.search(r"Quadra (.*?) - Lote", div.text, re.M)[1]
            lote = re.search(r"Lote (.*?)\s*Visualizar", div.text, re.M)[1]
            info = {
                "inscricao": inscricao,
                "bairro": bairro,
                "quadra": quadra,
                "lote": lote,
            }
            infos_lista.append(info)

        self._planilha(infos_lista)

    def _planilha(self, infos_lista: list[dict]) -> None:
        """
        Processa uma lista de informações e salva em uma planilha Excel.

        :param infos_lista: Lista de dicionários contendo informações de imóveis (inscricao, lote, quadra, bairro)
        :return: None
        """
        planilha_inicial = "path da planilha inicial"
        workbook = load_workbook(planilha_inicial)
        sheet = workbook["Página1"]
        linha_atual = 2

        for info in infos_lista:
            sheet.cell(row=linha_atual, column=3, value=info["inscricao"])
            sheet.cell(row=linha_atual, column=13, value=info["lote"])  # Lote
            sheet.cell(row=linha_atual, column=14, value=info["quadra"])  # Quadra
            sheet.cell(row=linha_atual, column=15, value=info["bairro"])  # Bairro
            linha_atual += 1

        planilha_resultado = "path da planilha resultado"
        workbook.save(planilha_resultado)
